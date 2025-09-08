# isms_report.py の内容 (修正後)

import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import os
# このファイルの親フォルダ（kikakuportal）をパスに追加
#sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import json
import time
from datetime import datetime
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.microsoft import EdgeChromiumDriverManager

from cryptography.fernet import Fernet
from dotenv import load_dotenv
from openpyxl import load_workbook


class HrmosAutomation:
    def __init__(self):
        self.today = None
        self.shokan = None
        self.driver = None
        self.wait = None
        self.jyukosha_name = None
        self.file_path = None
        self.executor_name = None # ISMS用にも追加
    def initialize(self):
        # (既存のコードと同じ。ただしprintデバッグを追加するならここに) 
        try:
            self.today = self.get_today()
            self.shokan = self.get_shokan()
            #self.driver, self.wait = self.setup_browser()
            self.setup_browser()
            self.file_path, self.jyukosha_name = self.process_excel()

            if self.file_path is None or self.jyukosha_name is None:
                messagebox.showerror("エラー", "Excelファイルからのデータ取得に失敗しました。処理を中断します。")
                if self.driver:
                    self.driver.quit()
                raise ValueError("Excelデータ不足のため処理を停止します。")

        except Exception as e:
            messagebox.showerror("初期化エラー", f"ISMS側の初期化中にエラーが発生しました:\n{e}")
            if self.driver:
                self.driver.quit()
            raise

    def get_today(self):
        return datetime.today().strftime('%Y/%m/%d')

    def get_shokan(self):
        kekka = messagebox.askquestion("確認", "確認テストは全問正解でしたか？")
        if kekka == 'yes':
           return "「ISMS 情報セキュリティ入場者研修テキスト」にて入場者研修を行い、 受講者に対して「情報セキュリティ意識テスト」による理解度を計った。今回の受講者において不正解箇所は無かった。セキュリティの意識・認識は得られたと感じている。"
        else:
           return "ＩＳＭＳ研修資料の新規従業員研修マニュアルにて入場者の研修を実施し情報セキュリティ意識テストによる本人の理解度を確認した。不正解箇所について、個別指導を行ったことで当社の情報セキュリティに関する認識は得られたと感じている。"
    
    def setup_browser(self):
        options = Options()
        options.add_argument("--guest")
        options.add_argument("--disable-features=EdgeIdentity")
        options.add_argument("--no-first-run")
        options.add_argument("--disable-popup-blocking")
        try:
            # msedgedriver.exe がスクリプトと同じディレクトリにある場合、またはパスが通っている場合
            msedgedriver_path = "msedgedriver.exe"
            service = Service(executable_path=msedgedriver_path)

            # ドライバーを初期化し、クラス変数 self.driver に格納
            self.driver = webdriver.Edge(service=service, options=options)

            # WebDriverWait も初期化し、クラス変数 self.wait に格納
            self.wait = WebDriverWait(self.driver, 15)

            print("Edgeブラウザが正常に起動・初期化されました。")
            # 起動成功後、例えば適当なページにアクセスしてみる
            # self.driver.get("https://www.google.com")
            # print(f"現在のタイトル: {self.driver.title}")
        except Exception as e:
            messagebox.showerror("ブラウザ起動エラー", f"Edgeブラウザの起動に失敗しました。WebDriverの確認やブラウザのインストール状況を確認してください。\nエラー: {e}")
            self.driver = None # 失敗時はNoneにリセット
            self.wait = None   # 失敗時はNoneにリセット
            raise

    def process_excel(self):
        ichiran = r"\\s106t001\share\100_共通\001_自由箱\新規入場者研修\テストパス一覧.xlsx"
        if not os.path.exists(ichiran):
            messagebox.showerror("ファイルエラー", f"ISMS: 指定されたExcelファイルが見つかりません:\n{ichiran}")
            return None, None

        try:
            df = pd.read_excel(ichiran, engine="openpyxl", header=None)
            wb = load_workbook(ichiran)
            ws = wb.active
            last_valid_index = df[df.columns[2]].last_valid_index()
            if last_valid_index is not None and last_valid_index >= 0:
                target_row_index = last_valid_index + 1
                if target_row_index < len(df):
                    file_path = df.iloc[target_row_index, 1]
                    jyukosha_name = df.iloc[target_row_index, 3]
                    file_path = os.path.normpath(str(file_path).replace("￥", "\\").replace("\u3000", "　").strip())
                    jyukosha_name = str(jyukosha_name).replace("\u3000", "　").strip()
                    ws.cell(row=target_row_index + 1, column=3).value = self.today
                    wb.save(ichiran)
                    return file_path, jyukosha_name
                else:
                    messagebox.showerror("Excelデータエラー", "ISMS: Excelファイルに次の処理データが見つかりません。")
                    return None, None
            else:
                messagebox.showerror("Excelデータエラー", "ISMS: Excelファイルに有効なデータがありません。")
                return None, None
        except Exception as e:
            messagebox.showerror("Excel処理エラー", f"ISMS: Excelファイルの処理中にエラーが発生しました:\n{e}")
            return None, None

    def open_login_page(self):
        script_dir = Path(__file__).parent
        dotenv_path = script_dir / '.env'
        if not dotenv_path.exists():
            messagebox.showerror("設定ファイルエラー", f"ISMS: .envファイルが見つかりません: {dotenv_path}")
            raise FileNotFoundError(".envファイルが見つかりません。")
        load_dotenv(dotenv_path=dotenv_path)

        key = os.getenv("SECRET_KEY")
        if not key:
            messagebox.showerror("設定エラー", "ISMS: .envファイルにSECRET_KEYが設定されていません。")
            raise ValueError("SECRET_KEYが.envファイルに設定されていません。")
        cipher = Fernet(key.encode())

        config_path = script_dir / 'config.json'
        if not config_path.exists():
            messagebox.showerror("設定ファイルエラー", f"ISMS: config.jsonファイルが見つかりません: {config_path}")
            raise FileNotFoundError("config.jsonファイルが見つかりません。")

        try:
            with open(config_path, "r", encoding="utf-8") as f:
                encrypted = json.load(f)
            username = cipher.decrypt(encrypted["username"].encode()).decode()
            password = cipher.decrypt(encrypted["password"].encode()).decode()
        except Exception as e:
            messagebox.showerror("認証情報エラー", f"ISMS: 認証情報の読み込みまたは復号化に失敗しました。config.jsonや.envファイルを確認してください。\nエラー: {e}")
            raise

        try:
            self.driver.get("https://hrmos-ursystems.ekeihi.net/")
            self.wait.until(EC.presence_of_element_located((By.ID, "LoginMenu_UserName"))).send_keys(username)
            self.wait.until(EC.presence_of_element_located((By.ID, "LoginMenu_Password"))).send_keys(password)
            self.wait.until(EC.element_to_be_clickable((By.ID, "LoginMenu_btnLogin"))).click()
            self.wait.until(EC.url_changes("https://hrmos-ursystems.ekeihi.net/"))
            self.driver.get("https://hrmos-ursystems.ekeihi.net/UI/Main/WorkflowInput.aspx?pageMode=1&parentPageDispID=&wfkbn=41&culture=1") # ISMS用のURL
        except Exception as e:
            messagebox.showerror("ログインエラー", f"ISMS: HRMOSへのログイン中にエラーが発生しました。\nエラー: {e}")
            raise

    def show_executor_selector(self):
        root = tk.Tk()
        root.title("実施者選択 (ISMS)")
        root.attributes("-topmost", True)

        executor_file_path = r"C:\Users\016215\source\repos\ISMSLauncher\ISMSLauncher\executors.csv"
        if not os.path.exists(executor_file_path):
            messagebox.showerror("ファイルエラー", f"ISMS: 実施者リストファイルが見つかりません:\n{executor_file_path}")
            root.destroy()
            return

        try:
            df_list = pd.read_csv(executor_file_path, encoding="utf-8-sig")
            executors = df_list.iloc[:, 0].tolist()
        except Exception as e:
            messagebox.showerror("ファイル読み込みエラー", f"ISMS: 実施者リストファイルの読み込みに失敗しました:\n{e}")
            root.destroy()
            return

        ttk.Label(root, text="ISMS実施者を選択：").pack(pady=5)
        combo = ttk.Combobox(root, values=executors, state="readonly")
        combo.pack(pady=5)
        if executors:
            combo.set(executors[0])

        def submit():
            selected = combo.get()
            if selected:
                self.executor_name = selected # self.executor_name を使用
                root.destroy()
            else:
                messagebox.showwarning("未選択", "実施者を選択してください。")

        combo.bind("<<ComboboxSelected>>", lambda event: submit())
        ttk.Button(root, text="決定", command=submit).pack(pady=10)
        root.wait_window() # root.mainloop() の代わりに wait_window() を使用

    def fill_form_and_upload(self):
        if not self.executor_name:
            self.show_executor_selector()
        if not self.executor_name: # 修正: global executor_name ではなく self.executor_name を確認
            messagebox.showwarning("処理中止", "実施者が選択されなかったため、処理を中止します。")
            return

        try:
            date_input = self.driver.find_element(By.ID, "ctrl_0130")
            date_input.clear()
            date_input.send_keys(self.today)
            
            executor_input = self.driver.find_element(By.ID, "ctrl_0112")
            executor_input.clear()
            executor_input.send_keys(self.jyukosha_name)
            
            executor_input = self.driver.find_element(By.ID, "ctrl_0127")
            executor_input.clear()
            executor_input.send_keys(self.executor_name) # self.executor_name を使用

            executor_input = self.driver.find_element(By.ID, "ctrl_0135")
            executor_input.clear()
            executor_input.send_keys("ISMS研修")

            executor_input = self.driver.find_element(By.ID, "ctrl_0138")
            executor_input.clear()
            executor_input.send_keys(self.shokan)

            self.wait.until(EC.element_to_be_clickable((By.ID, "cphBody_btnTmpRegister"))).click()

            link = self.wait.until(EC.presence_of_element_located((By.ID, "ctrl_0009_Lnk")))
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", link)
            self.driver.execute_script("arguments[0].click();", link)

            file_input = self.wait.until(EC.presence_of_element_located((By.ID, "cphBody_ucClipFileList_fupOldUpload")))
            file_input.send_keys(self.file_path)
            
            time.sleep(1) # 2秒間待機
            actions = ActionChains(self.driver)
            actions.move_by_offset(10, 10).click().perform()
            
        except Exception as e:
            messagebox.showerror("フォーム入力/アップロードエラー", f"ISMS: フォーム入力またはファイルアップロード中にエラーが発生しました:\n{e}")
            raise

    def run(self):
        try:
            self.initialize()
            self.open_login_page()
            self.fill_form_and_upload()
            messagebox.showinfo("完了", "ISMS報告処理が正常に完了しました。")
        except Exception as e:
            messagebox.showerror("ISMS自動化エラー", f"ISMS報告自動化処理中に予期せぬエラーが発生しました:\n{e}")
        finally:
            if self.driver:
                self.driver.quit()

# if __name__ == "__main__":
#     automation = HrmosAutomation()
#     automation.run()

