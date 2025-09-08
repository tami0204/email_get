# pms_report.py の内容 (前回の「完全版」をそのまま使用。末尾のif __name__ == "__main__":ブロックはコメントアウト)

import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import os
import json
import time
from datetime import datetime
from pathlib import Path

# Selenium関連のインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.microsoft import EdgeChromiumDriverManager

# 暗号化関連のインポート
from cryptography.fernet import Fernet
from dotenv import load_dotenv

# Excel関連のインポート
from openpyxl import load_workbook


class HrmosAutomationp:
    """
    PMS報告に関連するHrmosの自動化処理を行うクラス。
    ブラウザ操作、Excelデータの読み書き、フォーム入力、ファイルアップロードなどを行う。
    """
    def __init__(self):
        # クラス内で使用する変数を初期化
        self.today = None
        self.driver = None
        self.wait = None
        self.jyukosha_name = None
        self.file_path = None
        self.executor_name = None # 実施者名を保持するインスタンス変数

    def initialize(self):
        """
        自動化処理の初期設定（日付取得、ブラウザセットアップ、Excel処理）を行う。
        エラーが発生した場合は、メッセージボックスで通知して例外を発生させる。
        """
        try:
            self.today = self.get_today()
            self.setup_browser()
            self.file_path, self.jyukosha_name = self.process_excel()

            # Excel処理で必要な情報が取得できなかった場合
            if self.file_path is None or self.jyukosha_name is None:
                messagebox.showerror("エラー", "Excelファイルからのデータ取得に失敗しました。処理を中断します。")
                # driverを閉じ、エラーを再raiseしてプログラムを終了させる
                if self.driver:
                    self.driver.quit()
                raise ValueError("Excelデータ不足のため処理を停止します。")

        except Exception as e:
            messagebox.showerror("初期化エラー", f"PMS側の初期化中にエラーが発生しました:\n{e}")
            if self.driver:
                self.driver.quit()
            raise # 例外を再スローして、呼び出し元で捕捉させる

    def get_today(self):
        """現在の処理日を 'YYYY/MM/DD' 形式で取得する。"""
        return datetime.today().strftime('%Y/%m/%d')

    def setup_browser(self):
        """
        Microsoft Edgeブラウザを設定し、WebDriverとWebDriverWaitのインスタンスを返す。
        ゲストモードで起動し、Edge Identityやポップアップブロックを無効化する。
        """
        options = Options()
        options.add_argument("--guest")
        options.add_argument("--disable-features=EdgeIdentity")
        options.add_argument("--no-first-run")
        options.add_argument("--disable-popup-blocking")

        try:
            # msedgedriver.exe がスクリプトと同じディレクトリにある場合、またはパスが通っている場合
            msedgedriver_path = "msedgedriver.exe"
            service = Service(executable_path=msedgedriver_path)

            # ドライバーを初期化し、クラス変数 self.driver に格納
            self.driver = webdriver.Edge(service=service, options=options)

            # WebDriverWait も初期化し、クラス変数 self.wait に格納
            self.wait = WebDriverWait(self.driver, 15)

            print("Edgeブラウザが正常に起動・初期化されました。")
            # 起動成功後、例えば適当なページにアクセスしてみる
            # self.driver.get("https://www.google.com")
            # print(f"現在のタイトル: {self.driver.title}")
        except Exception as e:
            messagebox.showerror("ブラウザ起動エラー", f"Edgeブラウザの起動に失敗しました。WebDriverの確認やブラウザのインストール状況を確認してください。\nエラー: {e}")
            raise # 例外を再スロー

    def process_excel(self):
        """
        指定されたExcelファイルからデータを読み込み、講師名とファイルパスを取得する。
        また、取得した行に処理日を書き込み、Excelファイルを保存する。
        """
        # Excelファイルのパス (例: 共有フォルダ上のパス)
        ichiran = r"\\s106t001\share\100_共通\001_自由箱\新規入場者研修\テストパス一覧.xlsx"
        
        if not os.path.exists(ichiran):
            messagebox.showerror("ファイルエラー", f"PMS: 指定されたExcelファイルが見つかりません:\n{ichiran}")
            return None, None # ファイルが見つからない場合はNoneを返す

        try:
            df = pd.read_excel(ichiran, engine="openpyxl", header=None)
            wb = load_workbook(ichiran)
            ws = wb.active

            # 3列目（日付列）の最後に値がある行のインデックスを取得
            last_valid_index = df[df.columns[2]].last_valid_index()

            if last_valid_index is not None and last_valid_index >= 0:
                target_row_index = last_valid_index + 1

                if target_row_index < len(df):
                    file_path = df.iloc[target_row_index, 1]
                    jyukosha_name = df.iloc[target_row_index, 3]

                    file_path = os.path.normpath(str(file_path).replace("￥", "\\").replace("\u3000", "　").strip())
                    jyukosha_name = str(jyukosha_name).replace("\u3000", "　").strip()

                    ws.cell(row=target_row_index + 1, column=3).value = self.today
                    wb.save(ichiran)
                    
                    return file_path, jyukosha_name
                else:
                    messagebox.showerror("Excelデータエラー", "PMS: Excelファイルに次の処理データが見つかりません。")
                    return None, None
            else:
                messagebox.showerror("Excelデータエラー", "PMS: Excelファイルに有効なデータがありません。")
                return None, None

        except Exception as e:
            messagebox.showerror("Excel処理エラー", f"PMS: Excelファイルの処理中にエラーが発生しました:\n{e}")
            return None, None # エラー発生時もNoneを返す

    def open_login_page(self):
        """
        .envファイルとconfig.jsonから認証情報を読み込み、HRMOSにログインする。
        """
        script_dir = Path(__file__).parent

        dotenv_path = script_dir / '.env'
        if not dotenv_path.exists():
            messagebox.showerror("設定ファイルエラー", f"PMS: .envファイルが見つかりません: {dotenv_path}")
            raise FileNotFoundError(".envファイルが見つかりません。")
        load_dotenv(dotenv_path=dotenv_path)

        key = os.getenv("SECRET_KEY")
        if not key:
            messagebox.showerror("設定エラー", "PMS: .envファイルにSECRET_KEYが設定されていません。")
            raise ValueError("SECRET_KEYが.envファイルに設定されていません。")
        cipher = Fernet(key.encode())

        config_path = script_dir / 'config.json'
        if not config_path.exists():
            messagebox.showerror("設定ファイルエラー", f"PMS: config.jsonファイルが見つかりません: {config_path}")
            raise FileNotFoundError("config.jsonファイルが見つかりません。")

        try:
            with open(config_path, "r", encoding="utf-8") as f:
                encrypted = json.load(f)

            username = cipher.decrypt(encrypted["username"].encode()).decode()
            password = cipher.decrypt(encrypted["password"].encode()).decode()
        except Exception as e:
            messagebox.showerror("認証情報エラー", f"PMS: 認証情報の読み込みまたは復号化に失敗しました。config.jsonや.envファイルを確認してください。\nエラー: {e}")
            raise

        try:
            self.driver.get("https://hrmos-ursystems.ekeihi.net/")
            self.wait.until(EC.presence_of_element_located((By.ID, "LoginMenu_UserName"))).send_keys(username)
            self.wait.until(EC.presence_of_element_located((By.ID, "LoginMenu_Password"))).send_keys(password)
            self.wait.until(EC.element_to_be_clickable((By.ID, "LoginMenu_btnLogin"))).click()
            self.wait.until(EC.url_changes("https://hrmos-ursystems.ekeihi.net/"))
            self.driver.get("https://hrmos-ursystems.ekeihi.net/UI/Main/WorkflowInput.aspx?pageMode=1&parentPageDispID=&wfkbn=49&culture=1") # PMS用のURL
        except Exception as e:
            messagebox.showerror("ログインエラー", f"PMS: HRMOSへのログイン中にエラーが発生しました。\nエラー: {e}")
            raise

    def show_executor_selector(self):
        """
        Tkinterウィンドウを表示し、実施者（講師）をCSVファイルから読み込んで選択させる。
        選択された値は self.executor_name に格納される。
        """
        root = tk.Tk()
        root.title("実施者選択 (PMS)")
        root.attributes("-topmost", True)

        executor_file_path = r"C:\Users\016215\source\repos\ISMSLauncher\ISMSLauncher\executors.csv"
        
        if not os.path.exists(executor_file_path):
            messagebox.showerror("ファイルエラー", f"PMS: 実施者リストファイルが見つかりません:\n{executor_file_path}")
            root.destroy()
            return

        try:
            df_list = pd.read_csv(executor_file_path, encoding="utf-8-sig")
            executors = df_list.iloc[:, 0].tolist()
        except Exception as e:
            messagebox.showerror("ファイル読み込みエラー", f"PMS: 実施者リストファイルの読み込みに失敗しました:\n{e}")
            root.destroy()
            return

        ttk.Label(root, text="PMS実施者を選択：").pack(pady=5)
        combo = ttk.Combobox(root, values=executors, state="readonly")
        combo.pack(pady=5)
        
        if executors:
            combo.set(executors[0])

        def submit():
            selected = combo.get()
            if selected:
                self.executor_name = selected
                root.destroy()
            else:
                messagebox.showwarning("未選択", "実施者を選択してください。")

        combo.bind("<<ComboboxSelected>>", lambda event: submit())
        ttk.Button(root, text="決定", command=submit).pack(pady=10)
        root.wait_window() # root.mainloop() の代わりに wait_window() を使用

    def fill_form_and_upload(self):
        """
        HRMOSのフォームにデータを入力し、ファイルをアップロードする。
        """
        if not self.executor_name:
            self.show_executor_selector()
        if not self.executor_name:
            messagebox.showwarning("処理中止", "実施者が選択されなかったため、処理を中止します。")
            return

        try:
            # 件名
            self.wait.until(EC.presence_of_element_located((By.ID, "ctrl_0010"))).send_keys("個人情報保護にかかる新規入場者研修")
            # 作成日
            date_input = self.driver.find_element(By.ID, "ctrl_0102")
            date_input.clear()
            date_input.send_keys(self.today)
            #研修No
            date_input = self.driver.find_element(By.ID, "ctrl_0104")
            date_input.clear()
            date_input.send_keys("1")
            #講師
            executor_input = self.driver.find_element(By.ID, "ctrl_0109")
            executor_input.clear()
            executor_input.send_keys(self.executor_name)
            #実施場所
            executor_input = self.driver.find_element(By.ID, "ctrl_0111")
            executor_input.clear()
            executor_input.send_keys("自席")
            #使用するテキスト
            executor_input = self.driver.find_element(By.ID, "ctrl_0143")
            executor_input.clear()
            executor_input.send_keys("個人情報保護に関する研修資料")
            #研修予定日
            date_input = self.driver.find_element(By.ID, "ctrl_0120")
            date_input.clear()
            date_input.send_keys(self.today)

            # checkbox = self.driver.find_element(By.ID, "ctrl_0126")
            # if not checkbox.is_selected():
            #     checkbox.click()

            self.wait.until(EC.element_to_be_clickable((By.ID, "cphBody_btnTmpRegister"))).click()

            link = self.wait.until(EC.presence_of_element_located((By.ID, "ctrl_0009_Lnk")))
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", link)
            self.driver.execute_script("arguments[0].click();", link)

            file_input = self.wait.until(EC.presence_of_element_located((By.ID, "cphBody_ucClipFileList_fupOldUpload")))
            file_input.send_keys(self.file_path)
            
            time.sleep(1)
            actions = ActionChains(self.driver)
            actions.move_by_offset(10, 10).click().perform()
            
        except Exception as e:
            messagebox.showerror("フォーム入力/アップロードエラー", f"PMS: フォーム入力またはファイルアップロード中にエラーが発生しました:\n{e}")
            raise

    def runp(self):
        """
        PMS報告自動化処理のメイン実行メソッド。
        初期化、ログイン、フォーム入力・ファイルアップロードの各ステップを実行する。
        """
        try:
            self.initialize()
            self.open_login_page()
            self.fill_form_and_upload()
            messagebox.showinfo("完了", "PMS報告処理が正常に完了しました。")
        except Exception as e:
            messagebox.showerror("PMS自動化エラー", f"PMS報告自動化処理中に予期せぬエラーが発生しました:\n{e}")
        finally:
            if self.driver:
                # self.driver.quit()
                print("ドライバーは存在しますが、コメントアウトされているため終了しません。")


# このファイル（pms_report.py）を単体で実行した場合のテストコード
# if __name__ == "__main__":
#     try:
#         automation = HrmosAutomationp()
#         automation.runp()
#     except Exception as e:
#         print(f"PMS報告自動化スクリプトの単体実行中にエラーが発生しました: {e}")
#         input("Press any key to continue . . .")

